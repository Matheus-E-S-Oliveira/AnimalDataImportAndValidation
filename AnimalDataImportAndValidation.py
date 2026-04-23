from datetime import datetime
import pandas as pd
from sqlalchemy import create_engine, text
from pathlib import Path
import json
from openpyxl.styles import PatternFill, Font, Alignment

def read_file(path: str):
    """Lê um arquivo JSON e retorna seu conteúdo.

    Args:
        path (str): Caminho do arquivo JSON.

    Returns:
        list | dict: Conteúdo do JSON convertido para estrutura Python.

    Raises:
        Exception: Caso o formato do arquivo não seja suportado.
    """
    if str(path).endswith(".json"):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    else:
        raise Exception("Formato de arquivo não suportado")
    
def extract_animals(animal, extracted_animals, visited):
    """Extrai dados de animais de forma recursiva, incluindo pai e mãe.

    Essa função percorre a estrutura hierárquica do JSON e transforma
    os dados em formato plano (flat), evitando duplicidade através
    do controle por identificador.

    Args:
        animal (dict): Objeto do animal atual.
        extracted_animals (list): Lista acumuladora dos animais extraídos.
        visited (set): Conjunto de identificadores já processados.

    Returns:
        None
    """
    if not animal: return
    
    if visited is None: visited = set()
    
    identifier = get_identifier(animal)
    
    if not identifier:
        identifier = f"TEMP_{animal.get('Nome', 'SEM_NOME')}_{animal.get('DataNascimento', '')}"
    
    if identifier in visited: return

    visited.add(identifier)
    
    father = animal.get("Pai")
    mother = animal.get("Mae")
    
    current = {
        "RGN": animal.get("Rgn"),
        "RGD": animal.get("Rgd") or None,
        "nome": animal.get("Nome"),
        "sexo": animal.get("Sexo"),
        "dataNascimento": animal.get("DataNascimento"),
        "paiId": get_identifier(father),
        "maeId": get_identifier(mother),
    }
    
    extracted_animals.append(current)
    
    extract_animals(father, extracted_animals, visited)
    extract_animals(mother, extracted_animals, visited)

def get_identifier(animal):
    """Obtém o identificador único de um animal.

    A prioridade é:
        1. RGD
        2. RGN

    Args:
        animal (dict): Objeto do animal.

    Returns:
        str | None: Identificador do animal ou None se não existir.
    """
    if not animal:
        return None
    
    return animal.get("Rgd") or animal.get("Rgn")

def process_data(data):
    """Processa os dados brutos do JSON em um DataFrame estruturado.

    Percorre todos os animais e extrai suas informações,
    incluindo relações de parentesco.

    Args:
        data (list): Lista de animais no formato JSON.

    Returns:
        pandas.DataFrame: DataFrame contendo os dados estruturados.
    """
    extracted_animals = []
    visited = set()
        
    for animal in data:
        extract_animals(animal, extracted_animals, visited)
            
    df = pd.DataFrame(extracted_animals)
    return df

def clean_data(df):
    """Realiza limpeza e normalização dos dados.

    Etapas realizadas:
        - Criação do identificador (RGD ou RGN)
        - Remoção de registros sem identificador
        - Conversão e validação de datas
        - Remoção de datas inválidas
        - Normalização de nomes
        - Remoção de duplicidades

    Args:
        df (pandas.DataFrame): DataFrame bruto.

    Returns:
        pandas.DataFrame: DataFrame limpo e validado.
    """
    df["identificador"] = df["RGD"].fillna(df["RGN"])
    df = df[df["identificador"].notna()]
    
    df["dataNascimento"] = pd.to_datetime(df["dataNascimento"], errors="coerce")
    invalid_dates = df[df["dataNascimento"].isna()]

    if not invalid_dates.empty:
        print(f"[WARN] {len(invalid_dates)} animais com data inválida")

    df = df[df["dataNascimento"].notna()]
        
    df["nome"] = df["nome"].fillna("").str.strip().str.upper()
    df = df.drop_duplicates(subset=["identificador", "nome"])
    
    return df

def database_connection():
    """Cria a conexão com o banco de dados SQL Server.

    Utiliza SQLAlchemy com driver PyODBC.

    Returns:
        sqlalchemy.Engine: Engine de conexão com o banco.
    """
    engine = create_engine(
    "mssql+pyodbc://@PC-RERUM000210/AnimaisDB?driver=ODBC+Driver+17+for+SQL+Server&trusted_connection=yes&TrustServerCertificate=yes"
    )
    
    return engine
            
def export_report_to_excel(result):

    # ==============================
    # 1. Converter resultado
    # ==============================
    data = result.mappings().all()
    df = pd.DataFrame(data)

    if df.empty:
        print("Nenhuma divergência encontrada.")
        return
    

    # ==============================
    # 2. Renomear colunas
    # ==============================
    df = df.rename(columns={
        "Identificador": "ID",
        "nome": "Nome",
        "DataNascimento_Banco": "Data Nascimento (Banco)",
        "Sexo_Banco": "Sexo (Banco)",
        "DataNascimento_Importacao": "Data Nascimento (Importação)",
        "Sexo_Importacao": "Sexo (Importação)",
        "Divergencia_DataNascimento": "Divergência Data",
        "Divergencia_Sexo": "Divergência Sexo",
        "CamposDivergentes": "Campos Divergentes"
    })

    # ==============================
    # 3. Formatação
    # ==============================
    df["Data Nascimento (Banco)"] = pd.to_datetime(df["Data Nascimento (Banco)"]).dt.strftime("%d/%m/%Y")
    df["Data Nascimento (Importação)"] = pd.to_datetime(df["Data Nascimento (Importação)"]).dt.strftime("%d/%m/%Y")
    df["Campos Divergentes"] = df["Campos Divergentes"].str.replace(";", "", regex=False)

    # ==============================
    # 4. Estatísticas
    # ==============================
    total = len(df)
    diverg_data = df["Divergência Data"].sum()
    diverg_sexo = df["Divergência Sexo"].sum()

    resumo = pd.DataFrame({
        "Métrica": [
            "Total de Registros com Divergência",
            "Divergência de Data",
            "Divergência de Sexo"
        ],
        "Valor": [
            total,
            diverg_data,
            diverg_sexo
        ]
    })

    # Agrupamento por tipo de divergência
    estatisticas = df["Campos Divergentes"].value_counts().reset_index()
    estatisticas.columns = ["Tipo de Divergência", "Quantidade"]

    # ==============================
    # 5. Exportar Excel
    # ==============================
    output_dir = Path("Reports")
    output_dir.mkdir(exist_ok=True)

    data_atual = datetime.now().strftime("%Y-%m-%d_%H-%M")
    file_name = output_dir / f"Relatorio_Divergencias_{data_atual}.xlsx"

    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:

        # Aba detalhada
        df.to_excel(writer, index=False, sheet_name="Detalhado")

        # Aba resumo
        resumo.to_excel(writer, index=False, sheet_name="Resumo")

        # Aba estatísticas
        estatisticas.to_excel(writer, index=False, sheet_name="Estatísticas")

        workbook = writer.book

        # ==============================
        # 6. FORMATAÇÃO DETALHADO
        # ==============================
        ws = workbook["Detalhado"]
        ws.freeze_panes = "A2"
        
        headers = [cell.value for cell in ws[1]]
        col_map = {name: idx for idx, name in enumerate(headers)}
        
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            
        fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Destacar divergências
        for row in ws.iter_rows(min_row=2):

            diverg_data = row[col_map["Divergência Data"]].value
            diverg_sexo = row[col_map["Divergência Sexo"]].value

            # 🔴 Data
            if diverg_data == 1:
                row[col_map["Data Nascimento (Banco)"]].fill = fill_red
                row[col_map["Data Nascimento (Importação)"]].fill = fill_red

            # 🔴 Sexo
            if diverg_sexo == 1:
                row[col_map["Sexo (Banco)"]].fill = fill_red
                row[col_map["Sexo (Importação)"]].fill = fill_red

        auto_adjust_column_width(ws)
        
        # ==============================
        # 7. FORMATAÇÃO RESUMO
        # ==============================
        ws_resumo = workbook["Resumo"]
        auto_adjust_column_width(ws_resumo)
        ws_resumo.freeze_panes = "A2"

        for row in ws_resumo.iter_rows():
            for cell in row:
                cell.font = Font(bold=True) if cell.row == 1 else Font()
                cell.alignment = Alignment(horizontal="center")

        # ==============================
        # 8. FORMATAÇÃO ESTATÍSTICAS
        # ==============================
        ws_est = workbook["Estatísticas"]
        ws_est.freeze_panes = "A2"
        auto_adjust_column_width(ws_est)

        for row in ws_est.iter_rows():
            for cell in row:
                cell.font = Font(bold=True) if cell.row == 1 else Font()
                cell.alignment = Alignment(horizontal="center")

    print(f"Relatório gerado com sucesso: {file_name}")

def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2

# ================== MAIN ==================
def main():
    """Executa o pipeline completo de importação e validação.

    Fluxo:
        1. Leitura do arquivo JSON
        2. Processamento dos dados
        3. Limpeza e normalização
        4. Inserção na tabela de staging
        5. Execução da procedure de processamento
        6. Exibição de divergências (se houver)

    Returns:
        None
    """
    file = Path("C:/Users/marhe/OneDrive/Área de Trabalho/Tickets/Repository_Rerum/Ticket_16/ACNB-01-04-2026-11-06-43_-_inscrições__Expozebu_2026 - Divergente.json")
    engine = database_connection()
    
    data = read_file(file)
    
    df = process_data(data)
    df = clean_data(df)
    
    df["RGD"] = df["RGD"].where(pd.notna(df["RGD"]), None)
    df["RGN"] = df["RGN"].where(pd.notna(df["RGN"]), None)
    df["identificador"] = df["identificador"].astype(str)
    df = df[df["identificador"] != "nan"]

    df["nome"] = df["nome"].fillna("").str.strip().str.upper()
    df["sexo"] = df["sexo"].fillna("").str.strip().str.upper()
    
    with engine.begin() as conn:
        conn.execute(text("TRUNCATE TABLE Animal_Importacao"))
        
        df.to_sql("Animal_Importacao", conn, if_exists="append", index=False, chunksize=100, method="multi")
        
        result = conn.execute(text("EXEC dbo.sp_Processar_Importacao_Animais"))
        
        if result.returns_rows:
            export_report_to_excel(result)
            
    
if __name__ == "__main__":
    main()